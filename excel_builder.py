import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io

# Color fills
FILLS = {
    'A': PatternFill('solid', fgColor='FFD7D7'),  # red tint  = critical
    'B': PatternFill('solid', fgColor='FFF3CD'),  # amber tint = moderate
    'C': PatternFill('solid', fgColor='D4EDDA'),  # green tint = low value
}
HEADER_FILL = PatternFill('solid', fgColor='0F3460')
HEADER_FONT = Font(name='Arial', color='FFFFFF', bold=True, size=10)
BODY_FONT   = Font(name='Arial', size=9)
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')


def build_excel(results, summary, station_name):
    """
    Build a formatted Excel workbook with:
    Sheet 1: Full results table with ABC color coding
    Sheet 2: Summary statistics
    Sheet 3: ABC Class bar chart
    Returns bytes that can be sent as a file download.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Results ─────────────────────────────────
    ws = wb.active
    ws.title = f'{station_name} Results'

    headers = [
        'Part Name', 'ABC Class', 'Consumption/Mo', 'Reorder Point',
        'Safety Stock', 'Recommended Qty', 'EOQ', 'Turnover Rate',
        'Annual Value (EGP)', 'Stock Value (EGP)', 'Lead Time (d)',
    ]
    col_w = [38, 10, 16, 15, 14, 17, 8, 14, 20, 19, 14]

    for col, (h, w) in enumerate(zip(headers, col_w), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for ri, item in enumerate(results, start=2):
        abc = item.get('abc_class', 'C')
        row_data = [
            item['part_name'],
            abc,
            item['consumption_rate'],
            item['reorder_point'],
            item['safety_stock'],
            item['recommended_qty'],
            item['eoq'],
            item['turnover_rate'],
            item['annual_value'],
            item['total_stock_value'],
            item['lead_time_days'],
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill   = FILLS.get(abc, FILLS['C'])
            cell.font   = BODY_FONT
            cell.border = THIN
            cell.alignment = LEFT if col == 1 else CENTER
        ws.row_dimensions[ri].height = 16

    # ── Sheet 2: Summary ─────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    summary_rows = [
        ('Metric', 'Value'),
        ('Station / Depot', station_name),
        ('Total Parts Analyzed', summary['total_parts']),
        ('Class A Parts (Critical)', summary['class_a']),
        ('Class B Parts (Moderate)', summary['class_b']),
        ('Class C Parts (Low Value)', summary['class_c']),
        ('Total Recommended Stock Value (EGP)', f"{summary['total_stock_value']:,.0f}"),
        ('Average Turnover Rate', summary['avg_turnover']),
        ('Calculation Errors', summary['errors']),
    ]
    for ri2, (label, val) in enumerate(summary_rows, start=1):
        c1 = ws2.cell(row=ri2, column=1, value=label)
        c2 = ws2.cell(row=ri2, column=2, value=val)
        if ri2 == 1:
            c1.fill = c2.fill = HEADER_FILL
            c1.font = c2.font = HEADER_FONT
        else:
            c1.font = c2.font = BODY_FONT
        c1.border = c2.border = THIN
        c1.alignment = LEFT
        c2.alignment = CENTER

    # ── Sheet 3: ABC Chart ────────────────────────────────
    ws3 = wb.create_sheet('ABC Chart')
    ws3['A1'] = 'Class'
    ws3['B1'] = 'Count'
    ws3['A2'] = 'A - Critical'
    ws3['B2'] = summary['class_a']
    ws3['A3'] = 'B - Moderate'
    ws3['B3'] = summary['class_b']
    ws3['A4'] = 'C - Low Value'
    ws3['B4'] = summary['class_c']

    chart = BarChart()
    chart.title    = 'ABC Inventory Classification'
    chart.y_axis.title = 'Number of Parts'
    chart.x_axis.title = 'ABC Class'
    chart.style   = 10
    chart.width   = 15
    chart.height  = 10
    data   = Reference(ws3, min_col=2, min_row=1, max_row=4)
    cats   = Reference(ws3, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, 'D2')

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

